from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import time

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _autosize_and_style(worksheet):
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for column_cells in worksheet.columns:
        max_length = 12
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)) + 2, 45))
        worksheet.column_dimensions[column_letter].width = max_length


def _write_sheet(writer, sheet_name: str, frame: pd.DataFrame):
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    _autosize_and_style(writer.book[sheet_name])


def _shorten_text(text, max_len=12):
    """تقصير النصوص الطويلة"""
    text = str(text)
    if len(text) > max_len:
        return text[:max_len] + ".."
    return text


def _create_bar_chart(data, x_col, y_col, title):
    """إنشاء مخطط شريطي محسّن"""
    fig, ax = plt.subplots(figsize=(8, 4.5))
    
    # تقصير الأسماء
    x = data[x_col].astype(str).apply(lambda t: _shorten_text(t, 12)).tolist()
    y = data[y_col].tolist()
    
    # ألوان متدرجة
    colors = plt.cm.Blues(np.linspace(0.4, 0.9, len(x)))[::-1]
    bars = ax.bar(x, y, color=colors)
    
    # إضافة القيم
    max_y = max(y) if y else 1
    for bar, val in zip(bars, y):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max_y * 0.01,
               f'{val:,.0f}', ha='center', va='bottom', fontsize=7, rotation=0)
    
    ax.set_ylabel('المبيعات', fontsize=9)
    ax.set_title(title, fontsize=11, fontweight='bold')
    
    # تدوير النصوص 45 درجة لتجنب التداخل
    plt.xticks(rotation=45, ha='right', fontsize=7)
    plt.yticks(fontsize=8)
    
    # تنسيق محور Y
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format(int(x), ',')))
    
    plt.tight_layout()
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
    img_buffer.seek(0)
    plt.close()
    return img_buffer


def _create_pie_chart(data, x_col, y_col, title):
    """إنشاء مخطط دائري محسّن"""
    fig, ax = plt.subplots(figsize=(7, 4.5))
    
    data = data.sort_values(y_col, ascending=False)
    labels = data[x_col].astype(str).apply(lambda t: _shorten_text(t, 15)).tolist()
    values = data[y_col].tolist()
    total = sum(values)
    
    if total > 0:
        colors = plt.cm.Blues(np.linspace(0.3, 0.9, len(labels)))[::-1]
        
        def autopct_format(pct):
            return f'{pct:.1f}%' if pct > 3 else ''
        
        wedges, texts, autotexts = ax.pie(
            values, 
            labels=labels, 
            autopct=autopct_format,
            colors=colors,
            startangle=90,
            pctdistance=0.75,
            textprops={'fontsize': 7}
        )
        ax.set_title(title, fontsize=11, fontweight='bold')
        
        for text in texts:
            text.set_fontsize(7)
        for autotext in autotexts:
            autotext.set_fontsize(7)
            autotext.set_color('white')
    
    plt.tight_layout()
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
    img_buffer.seek(0)
    plt.close()
    return img_buffer


def _create_pareto_chart(data, title):
    """إنشاء مخطط باريتو محسّن"""
    fig, ax = plt.subplots(figsize=(8, 4.5))
    
    data = data.sort_values('الحالي', ascending=False)
    x = data['الاسم'].astype(str).apply(lambda t: _shorten_text(t, 10)).tolist()
    y = data['الحالي'].tolist()
    
    # أعمدة
    colors = plt.cm.Blues(np.linspace(0.4, 0.9, len(x)))[::-1]
    bars = ax.bar(x, y, color=colors, alpha=0.8)
    ax.set_ylabel('المبيعات', color='blue', fontsize=9)
    ax.tick_params(axis='y', labelcolor='blue')
    
    # النسبة التراكمية
    total = sum(y)
    cumsum = 0
    cum_percentages = []
    for val in y:
        cumsum += val
        cum_percentages.append((cumsum / total) * 100 if total > 0 else 0)
    
    ax2 = ax.twinx()
    ax2.plot(x, cum_percentages, color='red', marker='o', linewidth=2, markersize=5)
    ax2.axhline(y=80, color='red', linestyle='--', alpha=0.6, linewidth=1.5)
    ax2.text(len(x)-1, 83, '80%', color='red', fontsize=8, ha='right')
    ax2.set_ylabel('النسبة التراكمية %', color='red', fontsize=9)
    ax2.tick_params(axis='y', labelcolor='red')
    ax2.set_ylim(0, 105)
    
    ax.set_title(title, fontsize=11, fontweight='bold')
    
    # تدوير النصوص
    plt.xticks(rotation=45, ha='right', fontsize=7)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format(int(x), ',')))
    
    plt.tight_layout()
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
    img_buffer.seek(0)
    plt.close()
    return img_buffer


def _create_trend_chart(data, x_col, y_col, title):
    """إنشاء مخطط اتجاهات محسّن"""
    fig, ax = plt.subplots(figsize=(8, 4.5))
    
    x = data[x_col].astype(str).tolist()
    y = data[y_col].tolist()
    
    # أعمدة
    colors = plt.cm.Blues(np.linspace(0.4, 0.8, len(x)))
    ax.bar(x, y, color=colors, alpha=0.7, label='المبيعات')
    
    # خط الاتجاه
    ax.plot(x, y, color='red', marker='o', linewidth=2.5, markersize=8, label='خط الاتجاه')
    
    ax.set_ylabel('المبيعات', fontsize=9)
    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.legend(loc='upper left', fontsize=8)
    
    # تدوير النصوص
    plt.xticks(rotation=30, ha='right', fontsize=8)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format(int(x), ',')))
    
    plt.tight_layout()
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
    img_buffer.seek(0)
    plt.close()
    return img_buffer


def _create_chart_image_matplotlib(df: pd.DataFrame, chart_type: str, title: str, x_col: str = "الاسم", y_col: str = "الحالي") -> BytesIO | None:
    """إنشاء صورة للشارت باستخدام matplotlib."""
    data = df.head(12).copy()
    if data.empty or len(data) < 2:
        return None
    
    data = data.dropna(subset=[x_col, y_col])
    if data.empty:
        return None
    
    try:
        if chart_type == "bar":
            return _create_bar_chart(data, x_col, y_col, title)
        elif chart_type == "pie":
            return _create_pie_chart(data, x_col, y_col, title)
        elif chart_type == "pareto":
            return _create_pareto_chart(data, title)
        elif chart_type == "trend":
            return _create_trend_chart(data, x_col, y_col, title)
        else:
            return None
    except Exception as e:
        print(f"❌ خطأ في إنشاء المخطط: {e}")
        return None


def _add_image_to_sheet(worksheet, img_bytes: BytesIO, cell_position: str, width: int = 350, height: int = 230):
    try:
        img_bytes.seek(0)
        img = XLImage(img_bytes)
        img.width = width
        img.height = height
        worksheet.add_image(img, cell_position)
        return True
    except Exception as e:
        print(f"❌ خطأ في إضافة الصورة: {e}")
        return False


def _write_sheet_with_charts(writer, sheet_name: str, frame: pd.DataFrame, chart_title: str = None):
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    _autosize_and_style(worksheet)
    
    if not frame.empty and len(frame) > 1 and chart_title:
        try:
            table_cols = len(frame.columns)
            chart_start_col = table_cols + 2
            empty_col_letter = get_column_letter(table_cols + 1)
            worksheet.column_dimensions[empty_col_letter].width = 3
            
            header_cell = worksheet.cell(row=1, column=chart_start_col)
            header_cell.value = "📊 المخططات البيانية"
            header_cell.font = Font(color="1E3A5F", bold=True, size=14)
            header_cell.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(
                start_row=1,
                start_column=chart_start_col,
                end_row=1,
                end_column=chart_start_col + 2
            )
            
            chart_types = [
                ("bar", "📊 مخطط شريطي"),
                ("pie", "🧩 مخطط دائري"),
            ]
            
            current_row = 2
            
            for chart_type, chart_label in chart_types:
                try:
                    img_bytes = _create_chart_image_matplotlib(frame, chart_type, f"{chart_title} - {chart_label}")
                    
                    if img_bytes:
                        title_cell = worksheet.cell(row=current_row, column=chart_start_col)
                        title_cell.value = chart_label
                        title_cell.font = Font(color="1E3A5F", bold=True, size=10)
                        title_cell.alignment = Alignment(horizontal="center")
                        worksheet.merge_cells(
                            start_row=current_row,
                            start_column=chart_start_col,
                            end_row=current_row,
                            end_column=chart_start_col + 2
                        )
                        current_row += 1
                        
                        cell_pos = f"{get_column_letter(chart_start_col)}{current_row}"
                        if _add_image_to_sheet(worksheet, img_bytes, cell_pos, width=380, height=240):
                            current_row += 17
                        else:
                            error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                            error_cell.value = f"⚠️ فشل إدراج الصورة"
                            error_cell.font = Font(color="EF4444", size=9)
                            current_row += 2
                    else:
                        error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                        error_cell.value = f"⚠️ لا توجد بيانات كافية"
                        error_cell.font = Font(color="EF4444", size=9)
                        current_row += 2
                        
                except Exception as e:
                    error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                    error_cell.value = f"⚠️ خطأ: {str(e)[:30]}"
                    error_cell.font = Font(color="EF4444", size=9)
                    current_row += 2
            
            for col in range(chart_start_col, chart_start_col + 3):
                col_letter = get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = 38
                
        except Exception as e:
            error_cell = worksheet.cell(row=1, column=5)
            error_cell.value = f"⚠️ تعذر إضافة المخططات: {str(e)[:50]}"
            error_cell.font = Font(color="EF4444", size=10)


def _write_pareto_sheet(writer, sheet_name: str, frame: pd.DataFrame, chart_title: str = "تحليل باريتو"):
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    _autosize_and_style(worksheet)
    
    if not frame.empty and len(frame) > 1:
        try:
            table_cols = len(frame.columns)
            chart_start_col = table_cols + 2
            empty_col_letter = get_column_letter(table_cols + 1)
            worksheet.column_dimensions[empty_col_letter].width = 3
            
            header_cell = worksheet.cell(row=1, column=chart_start_col)
            header_cell.value = "📊 مخطط باريتو"
            header_cell.font = Font(color="1E3A5F", bold=True, size=14)
            header_cell.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(
                start_row=1,
                start_column=chart_start_col,
                end_row=1,
                end_column=chart_start_col + 2
            )
            
            current_row = 2
            
            try:
                img_bytes = _create_chart_image_matplotlib(frame, "pareto", chart_title)
                
                if img_bytes:
                    title_cell = worksheet.cell(row=current_row, column=chart_start_col)
                    title_cell.value = "📊 مخطط باريتو (قاعدة 80/20)"
                    title_cell.font = Font(color="1E3A5F", bold=True, size=10)
                    title_cell.alignment = Alignment(horizontal="center")
                    worksheet.merge_cells(
                        start_row=current_row,
                        start_column=chart_start_col,
                        end_row=current_row,
                        end_column=chart_start_col + 2
                    )
                    current_row += 1
                    
                    cell_pos = f"{get_column_letter(chart_start_col)}{current_row}"
                    if _add_image_to_sheet(worksheet, img_bytes, cell_pos, width=380, height=240):
                        current_row += 17
                    else:
                        error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                        error_cell.value = f"⚠️ فشل إدراج الصورة"
                        error_cell.font = Font(color="EF4444", size=9)
                        current_row += 2
                else:
                    error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                    error_cell.value = f"⚠️ لا توجد بيانات كافية"
                    error_cell.font = Font(color="EF4444", size=9)
                    current_row += 2
                    
            except Exception as e:
                error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                error_cell.value = f"⚠️ خطأ: {str(e)[:30]}"
                error_cell.font = Font(color="EF4444", size=9)
                current_row += 2
            
            for col in range(chart_start_col, chart_start_col + 3):
                col_letter = get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = 38
                
        except Exception as e:
            error_cell = worksheet.cell(row=1, column=5)
            error_cell.value = f"⚠️ تعذر إضافة المخطط: {str(e)[:50]}"
            error_cell.font = Font(color="EF4444", size=10)


def _write_trend_sheet(writer, sheet_name: str, frame: pd.DataFrame, chart_title: str = "تحليل الاتجاهات"):
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    _autosize_and_style(worksheet)
    
    if not frame.empty and len(frame) > 1:
        try:
            table_cols = len(frame.columns)
            chart_start_col = table_cols + 2
            empty_col_letter = get_column_letter(table_cols + 1)
            worksheet.column_dimensions[empty_col_letter].width = 3
            
            header_cell = worksheet.cell(row=1, column=chart_start_col)
            header_cell.value = "📊 مخطط الاتجاهات"
            header_cell.font = Font(color="1E3A5F", bold=True, size=14)
            header_cell.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(
                start_row=1,
                start_column=chart_start_col,
                end_row=1,
                end_column=chart_start_col + 2
            )
            
            current_row = 2
            
            try:
                x_col = frame.columns[0]
                y_col = frame.columns[1] if len(frame.columns) > 1 else frame.columns[0]
                
                img_bytes = _create_chart_image_matplotlib(
                    frame, "trend", chart_title, x_col=x_col, y_col=y_col
                )
                
                if img_bytes:
                    title_cell = worksheet.cell(row=current_row, column=chart_start_col)
                    title_cell.value = "📈 مخطط اتجاه المبيعات"
                    title_cell.font = Font(color="1E3A5F", bold=True, size=10)
                    title_cell.alignment = Alignment(horizontal="center")
                    worksheet.merge_cells(
                        start_row=current_row,
                        start_column=chart_start_col,
                        end_row=current_row,
                        end_column=chart_start_col + 2
                    )
                    current_row += 1
                    
                    cell_pos = f"{get_column_letter(chart_start_col)}{current_row}"
                    if _add_image_to_sheet(worksheet, img_bytes, cell_pos, width=380, height=240):
                        current_row += 17
                    else:
                        error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                        error_cell.value = f"⚠️ فشل إدراج الصورة"
                        error_cell.font = Font(color="EF4444", size=9)
                        current_row += 2
                else:
                    error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                    error_cell.value = f"⚠️ لا توجد بيانات كافية"
                    error_cell.font = Font(color="EF4444", size=9)
                    current_row += 2
                    
            except Exception as e:
                error_cell = worksheet.cell(row=current_row, column=chart_start_col)
                error_cell.value = f"⚠️ خطأ: {str(e)[:30]}"
                error_cell.font = Font(color="EF4444", size=9)
                current_row += 2
            
            for col in range(chart_start_col, chart_start_col + 3):
                col_letter = get_column_letter(col)
                worksheet.column_dimensions[col_letter].width = 38
                
        except Exception as e:
            error_cell = worksheet.cell(row=1, column=5)
            error_cell.value = f"⚠️ تعذر إضافة المخطط: {str(e)[:50]}"
            error_cell.font = Font(color="EF4444", size=10)


def export_to_excel(
    metrics: dict,
    customers: pd.DataFrame,
    products: pd.DataFrame,
    representatives: pd.DataFrame,
    branches: pd.DataFrame,
    insights: pd.DataFrame,
    pareto_customers: pd.DataFrame = None,
    pareto_products: pd.DataFrame = None,
    trend_data: pd.DataFrame = None,
) -> BytesIO:
    start = time.time()
    output = BytesIO()

    dashboard = pd.DataFrame(
        [
            {"المؤشر": "إجمالي المبيعات الحالية", "القيمة": metrics["current_total"]},
            {"المؤشر": "إجمالي المبيعات السابقة", "القيمة": metrics["previous_total"]},
            {"المؤشر": "إجمالي الفرق", "القيمة": metrics["difference"]},
            {"المؤشر": "نسبة النمو", "القيمة": metrics["growth"]},
            {"المؤشر": "عدد العملاء", "القيمة": metrics["customers_count"]},
            {"المؤشر": "عدد المنتجات", "القيمة": metrics["products_count"]},
            {"المؤشر": "عدد المناديب", "القيمة": metrics["representatives_count"]},
            {"المؤشر": "عدد الفروع", "القيمة": metrics["branches_count"]},
        ]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_sheet(writer, "Dashboard", dashboard)
        _write_sheet_with_charts(writer, "العملاء", customers, "تحليل العملاء")
        _write_sheet_with_charts(writer, "المنتجات", products, "تحليل المنتجات")
        _write_sheet_with_charts(writer, "المناديب", representatives, "تحليل المناديب")
        _write_sheet_with_charts(writer, "الفروع", branches, "تحليل الفروع")
        
        if pareto_customers is not None and not pareto_customers.empty:
            _write_pareto_sheet(writer, "باريتو العملاء", pareto_customers, "تحليل باريتو للعملاء")
        
        if pareto_products is not None and not pareto_products.empty:
            _write_pareto_sheet(writer, "باريتو المنتجات", pareto_products, "تحليل باريتو للمنتجات")
        
        if trend_data is not None and not trend_data.empty:
            _write_trend_sheet(writer, "تحليل الاتجاهات", trend_data, "تحليل اتجاهات المبيعات")
        
        _write_sheet(writer, "Insights", insights)

        for worksheet in writer.book.worksheets:
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1

    output.seek(0)
    print(f"⏱️ [excel_export] إجمالي التصدير: {time.time() - start:.2f} ثانية")
    return output